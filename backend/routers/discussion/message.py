# [UPDATE] backend/routers/discussion/message.py
import base64
import io
import json
import re
import shutil
import uuid
from datetime import datetime
from functools import partial
from pathlib import Path
from typing import Any, AsyncGenerator, Dict, List, Optional
import traceback
import threading
import asyncio
import zipfile

# Third-Party Imports
import fitz  # PyMuPDF
from docx import Document as DocxDocument
from fastapi import (
    APIRouter, BackgroundTasks, Depends, File, Form, HTTPException, Query,
    UploadFile, status, Body)
from fastapi.encoders import jsonable_encoder
from fastapi.responses import (FileResponse, HTMLResponse, JSONResponse,
                               PlainTextResponse, StreamingResponse, Response)
from lollms_client import (LollmsClient, LollmsDiscussion, LollmsMessage,
                           LollmsPersonality, MSG_TYPE)
from pydantic import BaseModel
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session
from ascii_colors import ASCIIColors, trace_exception
import markdown2
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pptx import Presentation
from pptx.util import Inches
from bs4 import BeautifulSoup

# Local Application Imports
from backend.config import APP_VERSION, SERVER_CONFIG
from backend.db import get_db
from backend.db.models.config import TTIBinding as DBTTIBinding
from backend.db.models.db_task import DBTask
from backend.db.models.personality import Personality as DBPersonality
from backend.db.models.user import (User as DBUser, UserMessageGrade,
                                     UserStarredDiscussion)
from backend.discussion import get_user_discussion, get_user_discussion_manager
from backend.routers.discussion.helpers import get_discussion_and_owner_for_request
from backend.models import (UserAuthDetails, ArtefactInfo, ContextStatusResponse,
                            DataZones, DiscussionBranchSwitchRequest,
                            DiscussionDataZoneUpdate, DiscussionExportRequest,
                            DiscussionImageUpdateResponse,
                            DiscussionInfo, DiscussionImportRequest,
                            DiscussionRagDatastoreUpdate, DiscussionSendRequest,
                            DiscussionTitleUpdate, ExportContextRequest,
                            ExportData, LoadArtefactRequest, ManualMessageCreate,
                            MessageCodeExportRequest, MessageContentUpdate,
                            MessageGradeUpdate, MessageOutput, MessageExportPayload,
                            MessageUpdateWithImages, TaskInfo,
                            UnloadArtefactRequest, ArtefactCreateManual, ArtefactUpdate)
from backend.session import (get_current_active_user,
                             get_current_db_user_from_token,
                             get_safe_store_instance,
                             get_user_discussion_assets_path,
                             get_user_lollms_client,
                             get_user_temp_uploads_path, user_sessions,
                             build_lollms_client_from_params)
from backend.task_manager import task_manager, Task
from backend.settings import settings
from backend.routers.files import (
    md2_to_html,              # markdown2 renderer with extras
    html_to_docx_bytes,       # HTML → python-docx mapper
    md_to_pdf_bytes,          # Markdown → PDF (images/tables/code/TOC)
    md_to_pptx_bytes,         # Markdown → PPTX (slides via ---)
    html_wrapper              # Wrap HTML body in a shell
)
from backend.tasks.image_generation_tasks import _generate_image_task

# safe_store is needed for RAG callbacks
try:
    import safe_store
except ImportError:
    safe_store = None

message_grade_lock = threading.Lock()

class RegenerateImageRequest(BaseModel):
    prompt: Optional[str] = None # Optional override, otherwise use stored

class ToggleImageRequest(BaseModel):
    active: Optional[bool] = None

def build_message_router(router: APIRouter):
    @router.put("/{discussion_id}/messages/{message_id}/grade", response_model=MessageOutput)
    async def grade_discussion_message(discussion_id: str, message_id: str, grade_update: MessageGradeUpdate, current_user: UserAuthDetails = Depends(get_current_active_user), db: Session = Depends(get_db)):
        username = current_user.username
        db_user = db.query(DBUser).filter(DBUser.username == username).one()
        discussion_obj, _, _, _ = await get_discussion_and_owner_for_request(discussion_id, current_user, db, 'interact')
        if not discussion_obj: raise HTTPException(status_code=404, detail="Discussion not found.")
        with message_grade_lock:
            grade = db.query(UserMessageGrade).filter_by(user_id=db_user.id, discussion_id=discussion_id, message_id=message_id).first()
            if grade: grade.grade += grade_update.change
            else:
                grade = UserMessageGrade(user_id=db_user.id, discussion_id=discussion_id, message_id=message_id, grade=grade_update.change)
                db.add(grade)
            db.commit()
            db.refresh(grade)
            current_grade = grade.grade
        branch = discussion_obj.get_branch(discussion_obj.active_branch_id)
        target_message = next((msg for msg in branch if msg.id == message_id), None)
        if not target_message: raise HTTPException(status_code=404, detail="Message not found in active branch.")
        full_image_refs = [ f"data:image/png;base64,{img}" for img in target_message.images or []]

        msg_metadata = target_message.metadata or {}
        return MessageOutput(
            id=target_message.id, sender=target_message.sender, sender_type=target_message.sender_type, content=target_message.content,
            parent_message_id=target_message.parent_id, binding_name=target_message.binding_name, model_name=target_message.model_name,
            token_count=target_message.tokens, sources=msg_metadata.get('sources'), events=msg_metadata.get('events'),
            image_references=full_image_refs, user_grade=current_grade, created_at=target_message.created_at,
            branch_id=discussion_obj.active_branch_id, branches=None
        )

    @router.put("/{discussion_id}/messages/{message_id}", response_model=MessageOutput)
    async def update_discussion_message(
        discussion_id: str,
        message_id: str,
        payload: MessageUpdateWithImages,
        current_user: UserAuthDetails = Depends(get_current_active_user),
        db: Session = Depends(get_db)
    ):
        username = current_user.username
        discussion_obj, _, _, _ = await get_discussion_and_owner_for_request(discussion_id, current_user, db, 'interact')
        if not discussion_obj:
            raise HTTPException(status_code=404, detail="Discussion not found.")
        
        target_message = discussion_obj.get_message(message_id)
        if not target_message:
            raise HTTPException(status_code=404, detail="Message not found in discussion.")

        target_message.content = payload.content

        # Handle image updates
        final_images_b64 = []
        new_images_b64 = []
        
        all_image_uris = (payload.kept_images_b64 or []) + (payload.new_images_b64 or [])
        
        for uri in payload.kept_images_b64 or []:
            try:
                if isinstance(uri, str) and ',' in uri:
                    _, encoded = uri.split(',', 1)
                    final_images_b64.append(encoded)
                else:
                    final_images_b64.append(uri)
            except ValueError:
                final_images_b64.append(uri)

        for uri in payload.new_images_b64 or []:
            try:
                if isinstance(uri, str) and ',' in uri:
                    _, encoded = uri.split(',', 1)
                    new_images_b64.append(encoded)
                    final_images_b64.append(encoded)
                else:
                    new_images_b64.append(uri)
                    final_images_b64.append(uri)
            except ValueError:
                new_images_b64.append(uri)
                final_images_b64.append(uri)
        
        # We need to preserve existing active state or reset if images changed significantly
        # Simplest approach: Rebuild images list. Add new images as new packs.
        
        # 1. Update main image list
        target_message.images = final_images_b64
        
        # 2. If new images were added, process them as packs
        if new_images_b64:
             for img in new_images_b64:
                 if hasattr(target_message, 'add_image_pack'):
                     target_message.add_image_pack([img], group_type="upload", active_by_default=True, title="User Upload")
        
        discussion_obj.commit()

        db_user = db.query(DBUser).filter(DBUser.username == username).one()
        grade = db.query(UserMessageGrade.grade).filter_by(user_id=db_user.id, discussion_id=discussion_id, message_id=message_id).scalar() or 0
        
        full_image_refs = [f"data:image/png;base64,{img}" for img in target_message.images or []]

        msg_metadata = target_message.metadata or {}
        return MessageOutput(
            id=target_message.id, sender=target_message.sender, sender_type=target_message.sender_type,
            content=target_message.content, parent_message_id=target_message.parent_id,
            binding_name=target_message.binding_name, model_name=target_message.model_name,
            token_count=target_message.tokens, sources=msg_metadata.get('sources'),
            events=msg_metadata.get('events'), image_references=full_image_refs,
            active_images=target_message.active_images, user_grade=grade,
            created_at=target_message.created_at, branch_id=discussion_obj.active_branch_id
        )

    @router.put("/{discussion_id}/messages/{message_id}/images/{image_index}/toggle", response_model=MessageOutput)
    async def toggle_message_image_activation(
        discussion_id: str,
        message_id: str,
        image_index: int,
        payload: ToggleImageRequest = Body(default=None),
        current_user: UserAuthDetails = Depends(get_current_active_user),
        db: Session = Depends(get_db)
    ):
        discussion_obj, _, _, _ = await get_discussion_and_owner_for_request(discussion_id, current_user, db, 'interact')
        
        msg = discussion_obj.get_message(message_id)
        if not msg:
            raise HTTPException(status_code=404, detail="Message not found.")
        
        try:
            # Pass explicit active state if provided
            active_arg = payload.active if payload else None
            # The client library method updates internal state
            msg.toggle_image_pack_activation(image_index, active=active_arg)
            
            # Persist changes
            discussion_obj.commit()
            
        except IndexError:
            raise HTTPException(status_code=404, detail="Image index out of bounds.")
        except Exception as e:
            trace_exception(e)
            raise HTTPException(status_code=500, detail="Failed to toggle image activation.")

        # Re-fetch for clean output
        db_user = db.query(DBUser).filter(DBUser.username == current_user.username).one()
        grade = db.query(UserMessageGrade.grade).filter_by(user_id=db_user.id, discussion_id=discussion_id, message_id=message_id).scalar() or 0
        
        full_image_refs = [f"data:image/png;base64,{img}" for img in msg.images or []]
        msg_metadata = msg.metadata or {}

        # Ensure we return the *updated* active_images list from the message object
        return MessageOutput(
            id=msg.id, sender=msg.sender, sender_type=msg.sender_type,
            content=msg.content, parent_message_id=msg.parent_id,
            binding_name=msg.binding_name, model_name=msg.model_name,
            token_count=msg.tokens, sources=msg_metadata.get('sources'),
            events=msg_metadata.get('events'), image_references=full_image_refs,
            active_images=msg.active_images, user_grade=grade,
            created_at=msg.created_at, branch_id=discussion_obj.active_branch_id
        )

    @router.post("/{discussion_id}/messages/{message_id}/images/{image_index}/regenerate", response_model=TaskInfo, status_code=202)
    async def regenerate_message_image(
        discussion_id: str,
        message_id: str,
        image_index: int,
        payload: RegenerateImageRequest,
        current_user: UserAuthDetails = Depends(get_current_active_user),
        db: Session = Depends(get_db)
    ):
        discussion_obj, _, _, _ = await get_discussion_and_owner_for_request(discussion_id, current_user, db, 'interact')
        
        msg = discussion_obj.get_message(message_id)
        if not msg:
            raise HTTPException(status_code=404, detail="Message not found.")

        # Find metadata for this image
        msg_metadata = msg.metadata or {}
        gen_infos = msg_metadata.get("generated_image_infos", [])
        
        target_info = next((info for info in gen_infos if info.get('index') == image_index), None)
        
        prompt_to_use = payload.prompt
        width_to_use = 1024
        height_to_use = 1024
        if not prompt_to_use and target_info:
            prompt_to_use = target_info.get('prompt')
            width_to_use = target_info.get('width', 1024)
            height_to_use = target_info.get('height', 1024)
            
        if not prompt_to_use:
             raise HTTPException(status_code=400, detail="No prompt found for this image, and none provided.")

        def _update_message_image_task(task: Task, username, discussion_id, message_id, image_index, prompt, width, height):
            task.log(f"Regenerating image {image_index} in message {message_id}...")
            task.set_progress(10)
            
            try:
                # Re-fetch discussion inside thread
                discussion = get_user_discussion(username, discussion_id)
                msg_obj = discussion.get_message(message_id)
                
                lc = build_lollms_client_from_params(username=username, load_llm=False, load_tti=True)
                if not lc.tti:
                     raise Exception("TTI binding not available.")
                
                task.set_progress(30)
                # Generate
                img_bytes = lc.tti.generate_image(prompt=prompt, width=width, height=height)
                if not img_bytes:
                    raise Exception("Generation returned empty data.")
                
                b64 = base64.b64encode(img_bytes).decode('utf-8')
                
                # Update message using the new method for adding packs
                if hasattr(msg_obj, 'add_image_pack'):
                    msg_obj.add_image_pack([b64], group_type="generated", active_by_default=True, title=prompt)
                
                new_index = len(msg_obj.images) - 1
                meta = msg_obj.metadata or {}
                
                # Update Infos
                infos = meta.get("generated_image_infos", [])
                infos.append({
                    'index': new_index, 
                    'prompt': prompt,
                    'width': width,
                    'height': height
                })
                msg_obj.set_metadata_item("generated_image_infos", infos, discussion)
                
                discussion.commit()
                
                task.set_progress(100)
                task.log("Image added to gallery and activated.")
                
                # Return full message to sync frontend
                full_image_refs = [f"data:image/png;base64,{img}" for img in msg_obj.images or []]
                return {
                    "status": "image_generated_in_message",
                    "discussion_id": discussion_id,
                    "new_message": {
                        "id": msg_obj.id, "sender": msg_obj.sender, "content": msg_obj.content,
                        "created_at": msg_obj.created_at.isoformat() if hasattr(msg_obj.created_at, 'isoformat') else str(msg_obj.created_at),
                        "parent_message_id": msg_obj.parent_id, "binding_name": msg_obj.binding_name, "model_name": msg_obj.model_name,
                        "token_count": msg_obj.tokens, "metadata": msg_obj.metadata, "image_references": full_image_refs,
                        "active_images": msg_obj.active_images, "sender_type": msg_obj.sender_type
                    }
                }

            except Exception as e:
                task.log(f"Regeneration failed: {e}", "ERROR")
                raise e

        db_task = task_manager.submit_task(
            name=f"Regenerate Image {image_index}",
            target=_update_message_image_task,
            args=(current_user.username, discussion_id, message_id, image_index, prompt_to_use, width_to_use, height_to_use),
            description=f"Regenerating image for prompt: '{prompt_to_use[:30]}...'",
            owner_username=current_user.username
        )
        return db_task

    @router.post("/{discussion_id}/messages", response_model=MessageOutput)
    async def add_manual_message(
        discussion_id: str,
        payload: ManualMessageCreate,
        current_user: UserAuthDetails = Depends(get_current_active_user),
        db: Session = Depends(get_db)
    ):
        username = current_user.username
        discussion_obj, _, _, _ = await get_discussion_and_owner_for_request(discussion_id, current_user, db, "interact")
        if not discussion_obj:
            raise HTTPException(status_code=404, detail="Discussion found.")

        if payload.sender_type not in ['user', 'assistant']:
            raise HTTPException(status_code=400, detail="Invalid sender_type.")

        sender_name = username
        if payload.sender_type == 'assistant':
            db_pers = db.query(DBPersonality).filter(DBPersonality.id == current_user.active_personality_id).first()
            sender_name = db_pers.name if db_pers else "assistant"

        try:
            new_message = discussion_obj.add_message(
                sender=sender_name,
                content=payload.content,
                parent_id=payload.parent_message_id,
                sender_type=payload.sender_type
            )
            discussion_obj.commit()
        except Exception as e:
            trace_exception(e)
            raise HTTPException(status_code=500, detail=f"Failed to add message to discussion: {e}")

        return MessageOutput(
            id=new_message.id,
            sender=new_message.sender,
            sender_type=new_message.sender_type,
            content=new_message.content,
            parent_message_id=new_message.parent_id,
            binding_name=new_message.binding_name,
            model_name=new_message.model_name,
            token_count=new_message.tokens,
            sources=[],
            events=[],
            image_references=[],
            user_grade=0,
            created_at=new_message.created_at,
            border_id=discussion_obj.active_branch_id
        )

    @router.delete("/{discussion_id}/messages/{message_id}", status_code=200)
    async def delete_discussion_message(discussion_id: str, message_id: str, current_user: UserAuthDetails = Depends(get_current_active_user), db: Session = Depends(get_db)):
        username = current_user.username
        discussion_obj, _, _, _ = await get_discussion_and_owner_for_request(discussion_id, current_user, db, 'interact')
        if not discussion_obj:
            raise HTTPException(status_code=404, detail="Discussion not found.")
        
        try:
            message_to_delete = discussion_obj.get_message(message_id)
            if not message_to_delete:
                raise ValueError("Message to be deleted not found in discussion.")
            
            parent_id = message_to_delete.parent_id
            
            discussion_obj.delete_branch(message_id)
            
            # After deleting, set the active branch to the parent of the deleted branch
            discussion_obj.switch_to_branch(parent_id)

            discussion_obj.commit()
        except ValueError as e:
            raise HTTPException(status_code=404, detail=str(e))
        
        db_user = db.query(DBUser).filter(DBUser.username == username).one()
        db.query(UserMessageGrade).filter_by(user_id=db_user.id, discussion_id=discussion_id, message_id=message_id).delete(synchronize_session=False)
        db.commit()
        return {"message": "Message and its branch deleted successfully."}


    @router.post("/{discussion_id}/messages/{message_id}/export")
    async def export_message(
        discussion_id: str,
        message_id: str,
        payload: MessageExportPayload,
        current_user: UserAuthDetails = Depends(get_current_active_user),
        db: Session = Depends(get_db)
    ):
        discussion, _, _, _ = await get_discussion_and_owner_for_request(discussion_id, current_user, db)
        message = discussion.get_message(message_id)

        if not message:
            raise HTTPException(status_code=404, detail="Message not found")

        export_format = payload.format.lower()

        # Normalize format name for setting key
        setting_key_format = 'markdown' if export_format == 'md' else export_format
        setting_key = f"export_to_{setting_key_format}_enabled"

        if not settings.get(setting_key, False):
            raise HTTPException(status_code=403, detail=f"Export to '{export_format}' is disabled by the administrator.")

        content = message.content
        filename = f"message_{message_id[:8]}.{export_format}"
        media_type = "application/octet-stream"
        file_content = b''
        
        # --- Collect Images for Slide Export ---
        # Prioritize 'slide_item' grouped images if present to export just the slideshow
        slide_images_b64 = []
        metadata = message.metadata or {}
        image_groups = (metadata.get('image_generation_groups', []) + metadata.get('image_groups', []))
        
        all_images = message.images or []
        
        # Collect indices of images that are part of a slide_item pack
        slide_indices = []
        for grp in image_groups:
            # FIX: Use dictionary access since metadata items are dicts after JSON deserialization
            grp_type = grp.get('type')
            if grp_type == 'slide_item':
                slide_indices.extend(grp.get('indices', []))
        
        if len(slide_indices) > 0:
             # Slideshow mode: only export slide images
             for idx in slide_indices:
                 if idx < len(all_images):
                     slide_images_b64.append(all_images[idx])
             # Clear text content to generate pure image presentation
             content = "" 
        else:
             # Standard mode: attach all images
             slide_images_b64 = all_images

        try:
            if export_format == 'txt':
                media_type = "text/plain"
                file_content = content.encode('utf-8')

            elif export_format in ['md', 'markdown']:
                media_type = "text/markdown"
                file_content = content.encode('utf-8')

            elif export_format == 'html':
                media_type = "text/html"
                html_content = md2_to_html(content)  # markdown2
                file_content = html_wrapper(html_content, title="Export")  # consistent HTML shell

            elif export_format == 'pdf':
                media_type = "application/pdf"
                file_content = md_to_pdf_bytes(content, extra_images=slide_images_b64)

            elif export_format == 'docx':
                media_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
                html_content = md2_to_html(content)  # markdown2
                file_content = html_to_docx_bytes(html_content)  # map HTML → Word

            elif export_format == 'pptx':
                media_type = "application/vnd.openxmlformats-officedocument.presentationml.presentation"
                # Pass collected images (slides or attachments) to PPTX generator
                file_content = md_to_pptx_bytes(content, extra_images=slide_images_b64)

            elif export_format == 'xlsx':
                media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                wb = Workbook()
                table_regex = re.compile(r'(\|.*\|(?:\r?\n|\r)?\|[-| :]*\|(?:\r?\n|\r)?(?:\|.*\|(?:\r?\n|\r)?)+)')
                tables = table_regex.findall(content)

                if tables:
                    for i, table_md in enumerate(tables):
                        ws = wb.create_sheet(title=f"Table {i+1}")
                        lines = [line.strip() for line in table_md.strip().split('\n')]
                        data_rows = [line for line in lines if not re.match(r'^\|[-| :]*\|$', line)]
                        for row_idx, line in enumerate(data_rows):
                            cells = [cell.strip() for cell in line.strip('|').split('|')]
                            for col_idx, cell_text in enumerate(cells):
                                ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell_text)
                        for col in ws.columns:
                            max_length = 0
                            column = get_column_letter(col[0].column)
                            for cell in col:
                                if cell.value and len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            ws.column_dimensions[column].width = (max_length + 2)
                    if wb.sheetnames[0] == 'Sheet':
                        wb.remove(wb['Sheet'])
                else:
                    ws = wb.active
                    ws.title = "Content"
                    ws['A1'] = content
                    ws.column_dimensions['A'].width = 100
                    ws['A1'].alignment = ws['A1'].alignment.copy(wrapText=True)

                bio = io.BytesIO(); wb.save(bio); file_content = bio.getvalue()

            elif export_format == 'epub':
                media_type = "application/epub+zip"
                html_content = md2_to_html(content)
                file_content = html_wrapper(html_content, title="Export")  # placeholder xhtml; replace with proper EPUB packaging

            elif export_format == 'odt':
                media_type = "application/vnd.oasis.opendocument.text"
                html_content = md2_to_html(content)
                file_content = html_wrapper(html_content, title="Export")  # placeholder

            elif export_format == 'rtf':
                media_type = "application/rtf"
                text = BeautifulSoup(md2_to_html(content), "html.parser").get_text()
                rtf = r"{\rtf1\ansi " + text.replace("\\", r"\\").replace("{", r"\{").replace("}", r"\}").replace("\n", r"\par ") + "}"
                file_content = rtf.encode("utf-8", errors="ignore")

            elif export_format in ['tex', 'latex']:
                media_type = "application/x-tex"
                text = BeautifulSoup(md2_to_html(content), "html.parser").get_text()
                def esc(t: str) -> str:
                    rep = {"\\": r"\textbackslash{}", "&": r"\&", "%": r"\%", "$": r"\$", "#": r"\#", "_": r"\_", "{": r"\{", "}": r"\}", "~": r"\textasciitilde{}", "^": r"\textasciicircum{}"}
                    for k, v in rep.items(): t = t.replace(k, v)
                    return t
                tex = "\\documentclass{article}\n\\usepackage[utf8]{inputenc}\n\\begin{document}\n" + esc(text) + "\n\\end{document}\n"
                file_content = tex.encode("utf-8")

            else:
                raise HTTPException(status_code=400, detail="Unsupported export format")

        except ImportError as e:
            raise HTTPException(status_code=501, detail=f"A required library for '{export_format}' export is missing: {e.name}")
        except Exception as e:
            trace_exception(e)
            raise HTTPException(status_code=500, detail=f"Failed to generate export file: {e}")

        headers = {'Content-Disposition': f'attachment; filename=\"{filename}\"'}
        return Response(content=file_content, media_type=media_type, headers=headers)
